"""Tests for classified output normalization and workbook export."""

from __future__ import annotations

import json
import os
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from config import AppConfig
from pipeline.fetch_events import UserTimeline
from pipeline.fetch_users import CandidateUser
from pipeline.classify_users import (
    ClassifiedJourney,
    _fallback_classification,
    _load_classifications_from_workbook,
    _normalize_response,
    classify_users,
    normalize_dropoff_point,
)
from pipeline.export_results import AnalysisMetadata, EXCEL_DATETIME_FORMAT, export_results
from pipeline.map_events import MappedJourney, _map_single_timeline


class ConfigTests(unittest.TestCase):
    def test_load_defaults_to_local_raw_and_cached(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            env_path = Path(tmp_dir) / ".env"
            env_path.write_text("", encoding="utf-8")

            with patch.dict(os.environ, {}, clear=True):
                config = AppConfig.load(env_path=env_path)

        self.assertIsNone(config.posthog_user_limit)
        self.assertEqual(config.data_source, "local_raw")
        self.assertEqual(config.classification_source, "cached")

    def test_blank_user_limit_loads_as_none(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            env_path = Path(tmp_dir) / ".env"
            env_path.write_text("POSTHOG_USER_LIMIT=\n", encoding="utf-8")

            with patch.dict(os.environ, {}, clear=True):
                config = AppConfig.load(env_path=env_path)

        self.assertIsNone(config.posthog_user_limit)

    def test_classification_source_openai_requires_openai_key(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            env_path = Path(tmp_dir) / ".env"
            env_path.write_text("CLASSIFICATION_SOURCE=openai\n", encoding="utf-8")

            with patch.dict(os.environ, {}, clear=True):
                config = AppConfig.load(env_path=env_path)

        with self.assertRaisesRegex(ValueError, "OPENAI_API_KEY"):
            config.validate()


class ClassificationTests(unittest.TestCase):
    def test_normalize_response_preserves_error_events(self) -> None:
        journey = _build_mapped_journey(
            error_events=["backend-errored-out", "network-error"],
            error_endpoint_urls=["https://api.focusbear.io/events"],
            error_status_codes=["403"],
            blocking_schedule_highest_stage="saved",
            error_event_occurrences=[
                {
                    "event": "backend-errored-out",
                    "endpoint_url": "https://api.focusbear.io/events",
                    "status_code": "403",
                    "count": 2,
                }
            ],
        )

        result = _normalize_response(
            journey,
            {
                "activated": False,
                "category": "Backend issue",
                "dropoff_point": "Routine Generated",
                "notes": "Backend failed after signup.",
                "pre_onboarding": "YES",
                "focus_bear_jr_greeting": "YES",
                "sign_up": "YES",
                "habits_introduction": "NO",
                "import_habits": "NO",
                "selecting_goals": "NO",
                "routine_generated": "NO",
                "blocking_intro": "NO",
                "screen_time_access": "NO",
                "set_up_blocking_schedule": "NO",
                "onboarding_complete": "NO",
                "home_screen": "NO",
            },
        )

        self.assertEqual(result.error_events, ["backend-errored-out", "network-error"])
        self.assertEqual(result.error_endpoint_urls, ["https://api.focusbear.io/events"])
        self.assertEqual(result.error_status_codes, ["403"])
        self.assertEqual(result.blocking_schedule_highest_stage, "saved")
        self.assertEqual(len(result.error_event_occurrences), 1)

    def test_fallback_classification_preserves_error_events(self) -> None:
        journey = _build_mapped_journey(
            error_events=["signin-error"],
            error_endpoint_urls=["https://api.focusbear.io/device"],
            error_status_codes=["400"],
            permission_events=["request-overlay-permissions"],
            blocking_schedule_highest_stage="configured",
            error_event_occurrences=[
                {
                    "event": "backend-errored-out",
                    "endpoint_url": "https://api.focusbear.io/device",
                    "status_code": "400",
                    "count": 1,
                }
            ],
        )

        result = _fallback_classification(journey, "OpenAI unavailable")

        self.assertEqual(result.category, "Backend issue")
        self.assertEqual(result.error_events, ["signin-error"])
        self.assertEqual(result.error_endpoint_urls, ["https://api.focusbear.io/device"])
        self.assertEqual(result.error_status_codes, ["400"])
        self.assertEqual(result.blocking_schedule_highest_stage, "configured")
        self.assertEqual(len(result.error_event_occurrences), 1)

    def test_normalize_dropoff_point_collapses_stage_variants(self) -> None:
        self.assertEqual(normalize_dropoff_point("set_up_blocking_schedule"), "Set Up Blocking Schedule")
        self.assertEqual(normalize_dropoff_point("Set Up Blocking Schedule"), "Set Up Blocking Schedule")
        self.assertEqual(normalize_dropoff_point("screen-time-access"), "Screen Time Access")
        self.assertEqual(normalize_dropoff_point("something custom"), "Unknown")


class ExportResultsTests(unittest.TestCase):
    def test_export_results_writes_readable_workbook(self) -> None:
        rows = [
            _build_classified_journey(
                user_id="user-backend",
                first_app_opened_at="2026-06-01T00:15:00+00:00",
                last_event_at="2026-06-01T03:45:00+00:00",
                category="Backend issue",
                error_events=["backend-errored-out", "network-error"],
                error_endpoint_urls=[
                    "https://api.focusbear.io/events",
                    "https://api.focusbear.io/blocking-schedules",
                ],
                error_status_codes=["403", "413"],
                blocking_schedule_highest_stage="saved",
                error_event_occurrences=[
                    {
                        "event": "backend-errored-out",
                        "endpoint_url": "https://api.focusbear.io/events",
                        "status_code": "403",
                        "count": 2,
                    },
                    {
                        "event": "network-error",
                        "endpoint_url": "https://api.focusbear.io/blocking-schedules",
                        "status_code": "",
                        "count": 1,
                    },
                ],
                notes="Two backend errors observed during onboarding.",
                pre_onboarding="YES",
                sign_up="YES",
                dropoff_point="set_up_blocking_schedule",
            ),
            _build_classified_journey(
                user_id="user-permission",
                first_app_opened_at="2026-06-02T01:00:00+00:00",
                last_event_at="2026-06-02T01:30:00+00:00",
                category="Permission issue",
                error_events=["backend-errored-out"],
                error_endpoint_urls=["https://api.focusbear.io/events"],
                error_status_codes=["400"],
                blocking_schedule_highest_stage="configured",
                error_event_occurrences=[
                    {
                        "event": "backend-errored-out",
                        "endpoint_url": "https://api.focusbear.io/events",
                        "status_code": "400",
                        "count": 1,
                    }
                ],
                notes="Permission prompt shown repeatedly.",
                dropoff_point="Set Up Blocking Schedule",
                onboarding_complete="YES",
            ),
            _build_classified_journey(
                user_id="user-early",
                first_app_opened_at="2026-06-03T01:00:00+00:00",
                last_event_at="2026-06-03T01:30:00+00:00",
                category="Early drop",
                error_events=[],
                error_endpoint_urls=[],
                error_status_codes=[],
                blocking_schedule_highest_stage="not_reached",
                error_event_occurrences=[],
                notes="User stopped before routine generation.",
                dropoff_point="Routine Generated",
            ),
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "onboarding_analysis.xlsx"
            export_results(
                rows,
                output_path,
                metadata=AnalysisMetadata(
                    cohort_id="239235",
                    cohort_name="People who didn't activate",
                    cohort_total_count=111,
                    analyzed_user_count=3,
                    posthog_user_limit=None,
                    lookback_days=90,
                    generated_at=datetime(2026, 6, 4, 2, 0, 0),
                ),
            )

            workbook = load_workbook(output_path)
            worksheet = workbook["Onboarding Analysis"]
            summary_sheet = workbook["Summary"]

            self.assertEqual(
                [cell.value for cell in worksheet[1]],
                [
                    "User ID",
                    "First App Opened At",
                    "Last Event At",
                    "Journey Duration",
                    "Category",
                    "Dropoff Point",
                    "Error Events",
                    "Error Endpoint URLs",
                    "Error Status Codes",
                    "Blocking Schedule Highest Stage",
                    "Notes",
                    "Pre Onboarding",
                    "Focus Bear Jr Greeting",
                    "Sign Up",
                    "Habits Introduction",
                    "Import Habits",
                    "Selecting Goals",
                    "Routine Generated",
                    "Blocking Intro",
                    "Screen Time Access",
                    "Set Up Blocking Schedule",
                    "Onboarding Complete",
                    "Home Screen",
                    "Raw Event Count",
                ],
            )
            self.assertEqual(workbook.sheetnames, ["Onboarding Analysis", "Summary"])
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(worksheet.auto_filter.ref, "A1:X4")

            self.assertEqual(worksheet["B2"].value, datetime(2026, 6, 1, 10, 15))
            self.assertEqual(worksheet["C2"].value, datetime(2026, 6, 1, 13, 45))
            self.assertEqual(worksheet["B2"].number_format, EXCEL_DATETIME_FORMAT)
            self.assertEqual(worksheet["C2"].number_format, EXCEL_DATETIME_FORMAT)

            self.assertEqual(worksheet["F2"].value, "Set Up Blocking Schedule")
            self.assertEqual(worksheet["F3"].value, "Set Up Blocking Schedule")
            self.assertEqual(worksheet["G2"].value, "backend-errored-out, network-error")
            self.assertEqual(worksheet["G3"].value, None)
            self.assertEqual(worksheet["G4"].value, None)
            self.assertEqual(
                worksheet["H2"].value,
                "https://api.focusbear.io/events, https://api.focusbear.io/blocking-schedules",
            )
            self.assertEqual(worksheet["H3"].value, "https://api.focusbear.io/events")
            self.assertEqual(worksheet["I2"].value, "403, 413")
            self.assertEqual(worksheet["I3"].value, "400")
            self.assertEqual(worksheet["J2"].value, "saved")
            self.assertEqual(worksheet["J3"].value, "configured")

            self.assertTrue(worksheet["G2"].alignment.wrap_text)
            self.assertTrue(worksheet["H2"].alignment.wrap_text)
            self.assertTrue(worksheet["I2"].alignment.wrap_text)
            self.assertTrue(worksheet["J2"].alignment.wrap_text)
            self.assertEqual(worksheet["G2"].alignment.vertical, "top")
            self.assertEqual(worksheet["L2"].alignment.horizontal, "center")

            self.assertTrue(_rgb(worksheet["A1"]).endswith("1F2937"))
            self.assertTrue(_rgb(worksheet["A2"]).endswith("F7F7F7"))
            self.assertFalse(_rgb(worksheet["A3"]).endswith("F7F7F7"))
            self.assertTrue(_rgb(worksheet["E2"]).endswith("F4CCCC"))
            self.assertTrue(_rgb(worksheet["E3"]).endswith("FFD966"))
            self.assertTrue(_rgb(worksheet["L2"]).endswith("C6EFCE"))
            self.assertTrue(_rgb(worksheet["M2"]).endswith("FFC7CE"))

            self.assertEqual(summary_sheet["A1"].value, "Metric")
            self.assertEqual(summary_sheet["A2"].value, "Generated At")
            self.assertEqual(summary_sheet["B2"].value, datetime(2026, 6, 4, 2, 0))
            self.assertEqual(summary_sheet["A5"].value, "Cohort Users Available")
            self.assertEqual(summary_sheet["B5"].value, 111)
            self.assertEqual(summary_sheet["A6"].value, "Users Analyzed")
            self.assertEqual(summary_sheet["B6"].value, 3)
            self.assertEqual(summary_sheet["A7"].value, "Applied User Limit")
            self.assertEqual(summary_sheet["B7"].value, "All cohort users")
            self.assertEqual(summary_sheet["A9"].value, "Onboarding Completed")
            self.assertEqual(summary_sheet["B9"].value, 1)
            self.assertEqual(summary_sheet["A10"].value, "Onboarding Completed %")
            self.assertEqual(summary_sheet["B10"].value, "33.3%")

            summary_rows = list(summary_sheet.iter_rows(values_only=True))
            self.assertIn(("Backend issue", 1, "33.3%", None, None, None), summary_rows)
            self.assertIn(("Permission issue", 1, "33.3%", None, None, None), summary_rows)
            self.assertIn(("Early drop", 1, "33.3%", None, None, None), summary_rows)
            self.assertIn(("Set Up Blocking Schedule", 2, None, None, None, None), summary_rows)
            self.assertIn(("Routine Generated", 1, None, None, None, None), summary_rows)
            self.assertIn(("saved", 1, "33.3%", None, None, None), summary_rows)
            self.assertIn(("configured", 1, "33.3%", None, None, None), summary_rows)
            self.assertIn(("backend-errored-out", 3, 2, "66.7%", None, None), summary_rows)
            self.assertIn(("network-error", 1, 1, "33.3%", None, None), summary_rows)
            self.assertIn(
                ("backend-errored-out", "https://api.focusbear.io/events", "403", 2, 1, "33.3%"),
                summary_rows,
            )
            self.assertIn(
                ("network-error", "https://api.focusbear.io/blocking-schedules", "(missing)", 1, 1, "33.3%"),
                summary_rows,
            )
            self.assertEqual(len(summary_sheet._charts), 0)

            findings = [row[0] for row in summary_rows if row and isinstance(row[0], str)]
            self.assertIn("Largest category: Permission issue (1/3, 33.3%).", findings)
            self.assertIn("Most common dropoff point: Set Up Blocking Schedule (2 users).", findings)
            self.assertIn(
                "Most common canonical error event: backend-errored-out (3 raw events across 2 users).",
                findings,
            )
            self.assertIn(
                "Most affected error breakdown: backend-errored-out @ https://api.focusbear.io/events [403] (1 users).",
                findings,
            )
            self.assertIn("Most common blocking-schedule deepest stage: saved (1 users).", findings)
            self.assertIn("Onboarding completion rate: 33.3% (1/3).", findings)

    def test_export_results_excludes_non_api_endpoints_from_detail_and_summary(self) -> None:
        timeline = _build_user_timeline(
            [
                _event("user-open-the-app-for-the-first-time", "2026-06-01T00:00:00+00:00"),
                _event(
                    "backend-errored-out",
                    "2026-06-01T00:01:00+00:00",
                    endpoint_url="https://events.aws.focusbear.io/events",
                    status_code="403",
                ),
                _event(
                    "network-error",
                    "2026-06-01T00:02:00+00:00",
                    endpoint_url="https://api.focusbear.io/events",
                ),
            ]
        )
        classified_row = _fallback_classification(_map_single_timeline(timeline), "test fallback")

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "onboarding_analysis.xlsx"
            export_results(
                [classified_row],
                output_path,
                metadata=AnalysisMetadata(
                    cohort_id="239235",
                    cohort_name="People who didn't activate",
                    cohort_total_count=1,
                    analyzed_user_count=1,
                    posthog_user_limit=None,
                    lookback_days=90,
                    generated_at=datetime(2026, 6, 4, 2, 0, 0),
                ),
            )

            workbook = load_workbook(output_path)
            worksheet = workbook["Onboarding Analysis"]
            summary_sheet = workbook["Summary"]

            self.assertEqual(worksheet["H2"].value, "https://api.focusbear.io/events")
            self.assertEqual(worksheet["I2"].value, None)

            summary_rows = list(summary_sheet.iter_rows(values_only=True))
            flattened = [str(value) for row in summary_rows for value in row if value is not None]
            self.assertFalse(any("events.aws.focusbear.io" in value for value in flattened))


class ClassificationCacheTests(unittest.TestCase):
    def test_classify_users_cached_loads_json_cache(self) -> None:
        journey = _build_mapped_journey()

        with tempfile.TemporaryDirectory() as tmp_dir:
            env_path = Path(tmp_dir) / ".env"
            env_path.write_text(
                "\n".join(
                    [
                        "DATA_SOURCE=local_raw",
                        "CLASSIFICATION_SOURCE=cached",
                        "OUTPUT_XLSX_PATH=data/outputs/onboarding_analysis.xlsx",
                    ]
                ),
                encoding="utf-8",
            )
            with patch.dict(os.environ, {}, clear=True):
                config = AppConfig.load(env_path=env_path)
            config.classified_journeys_cache_path = Path(tmp_dir) / "classified_journeys.json"

            cache_payload = [
                {
                    "user_id": "user-123",
                    "category": "Backend issue",
                    "dropoff_point": "Routine Generated",
                    "notes": "Cached classification",
                    "activated": False,
                    "pre_onboarding": "YES",
                    "focus_bear_jr_greeting": "YES",
                    "sign_up": "YES",
                    "habits_introduction": "NO",
                    "import_habits": "NO",
                    "selecting_goals": "NO",
                    "routine_generated": "NO",
                    "blocking_intro": "NO",
                    "screen_time_access": "NO",
                    "set_up_blocking_schedule": "NO",
                    "onboarding_complete": "NO",
                    "home_screen": "NO",
                }
            ]
            config.classified_journeys_cache_path.parent.mkdir(parents=True, exist_ok=True)
            config.classified_journeys_cache_path.write_text(
                json.dumps(cache_payload),
                encoding="utf-8",
            )

            results = classify_users(config=config, journeys=[journey], openai_client=None)

        self.assertEqual(len(results), 1)
        self.assertEqual(results[0].notes, "Cached classification")
        self.assertEqual(results[0].error_events, journey.error_events)

    def test_load_classifications_from_workbook_reuses_workbook_labels(self) -> None:
        journey = _build_mapped_journey(
            error_events=["backend-errored-out"],
            error_endpoint_urls=["https://api.focusbear.io/events"],
            error_status_codes=["403"],
            error_event_occurrences=[
                {
                    "event": "backend-errored-out",
                    "endpoint_url": "https://api.focusbear.io/events",
                    "status_code": "403",
                    "count": 2,
                }
            ],
        )
        workbook_row = _build_classified_journey(
            user_id="user-123",
            first_app_opened_at="2026-06-01T00:15:00+00:00",
            last_event_at="2026-06-01T01:15:00+00:00",
            category="Permission issue",
            error_events=[],
            error_endpoint_urls=[],
            error_status_codes=[],
            blocking_schedule_highest_stage="not_reached",
            error_event_occurrences=[],
            notes="Workbook classification",
            pre_onboarding="YES",
            sign_up="YES",
        )

        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "onboarding_analysis.xlsx"
            export_results(
                [workbook_row],
                workbook_path,
                metadata=AnalysisMetadata(
                    cohort_id="239235",
                    cohort_name="People who didn't activate",
                    cohort_total_count=1,
                    analyzed_user_count=1,
                    posthog_user_limit=None,
                    lookback_days=90,
                    generated_at=datetime(2026, 6, 4, 2, 0, 0),
                ),
            )

            results = _load_classifications_from_workbook(workbook_path, [journey])

        self.assertEqual(results[0].category, "Permission issue")
        self.assertEqual(results[0].notes, "Workbook classification")
        self.assertEqual(results[0].error_event_occurrences, journey.error_event_occurrences)


class MappingTests(unittest.TestCase):
    def test_map_single_timeline_extracts_error_endpoint_urls(self) -> None:
        timeline = _build_user_timeline(
            [
                _event("user-open-the-app-for-the-first-time", "2026-06-01T00:00:00+00:00"),
                _event(
                    "network-error",
                    "2026-06-01T00:01:00+00:00",
                    endpoint_url="https://api.focusbear.io/events",
                ),
                _event(
                    "backend-errored-out",
                    "2026-06-01T00:02:00+00:00",
                    endpoint_url="https://api.focusbear.io/blocking-schedules",
                    status_code="403",
                ),
                _event(
                    "backend-errored-out",
                    "2026-06-01T00:03:00+00:00",
                    endpoint_url="https://api.focusbear.io/events",
                    status_code="413",
                ),
                _event(
                    "backend-timed-out",
                    "2026-06-01T00:04:00+00:00",
                    endpoint_url="https://api.focusbear.io/ignored",
                ),
                _event(
                    "backend-errored-out",
                    "2026-06-01T00:05:00+00:00",
                    endpoint_url="https://events.aws.focusbear.io/events",
                    status_code="403",
                ),
            ]
        )

        journey = _map_single_timeline(timeline)

        self.assertEqual(journey.error_events, ["backend-errored-out", "backend-timed-out", "network-error"])
        self.assertEqual(
            journey.error_endpoint_urls,
            [
                "https://api.focusbear.io/blocking-schedules",
                "https://api.focusbear.io/events",
                "https://api.focusbear.io/ignored",
            ],
        )
        self.assertEqual(journey.error_status_codes, ["403", "413"])
        self.assertEqual(
            journey.error_event_occurrences,
            [
                {
                    "event": "backend-errored-out",
                    "endpoint_url": "https://api.focusbear.io/blocking-schedules",
                    "status_code": "403",
                    "count": 1,
                },
                {
                    "event": "backend-errored-out",
                    "endpoint_url": "https://api.focusbear.io/events",
                    "status_code": "413",
                    "count": 1,
                },
                {
                    "event": "backend-timed-out",
                    "endpoint_url": "https://api.focusbear.io/ignored",
                    "status_code": "",
                    "count": 1,
                },
                {
                    "event": "network-error",
                    "endpoint_url": "https://api.focusbear.io/events",
                    "status_code": "",
                    "count": 1,
                },
            ],
        )

    def test_map_single_timeline_filters_non_api_endpoints_out_of_error_signals(self) -> None:
        timeline = _build_user_timeline(
            [
                _event("user-open-the-app-for-the-first-time", "2026-06-01T00:00:00+00:00"),
                _event(
                    "backend-errored-out",
                    "2026-06-01T00:01:00+00:00",
                    endpoint_url="https://events.aws.focusbear.io/events",
                    status_code="403",
                ),
                _event(
                    "network-error",
                    "2026-06-01T00:02:00+00:00",
                    endpoint_url="https://events.aws.focusbear.io/events",
                ),
                _event("signin-error", "2026-06-01T00:03:00+00:00"),
            ]
        )

        journey = _map_single_timeline(timeline)

        self.assertEqual(journey.error_events, ["signin-error"])
        self.assertEqual(journey.error_endpoint_urls, [])
        self.assertEqual(journey.error_status_codes, [])
        self.assertEqual(journey.error_event_occurrences, [])

    def test_map_single_timeline_derives_blocking_schedule_highest_stage(self) -> None:
        cases = [
            (
                "opened",
                [
                    _event("user-open-the-app-for-the-first-time", "2026-06-01T00:00:00+00:00"),
                    _event("blocking-schedule-screen-opened", "2026-06-01T00:01:00+00:00", mode="edit"),
                ],
                "blocking-schedule-screen-opened",
            ),
            (
                "configured",
                [
                    _event("user-open-the-app-for-the-first-time", "2026-06-01T00:00:00+00:00"),
                    _event("blocking-schedule-add-new", "2026-06-01T00:01:00+00:00"),
                    _event(
                        "blocking-schedule-toggle-global",
                        "2026-06-01T00:02:00+00:00",
                        enabled=False,
                    ),
                ],
                "blocking-schedule-toggle-global",
            ),
            (
                "saved",
                [
                    _event("user-open-the-app-for-the-first-time", "2026-06-01T00:00:00+00:00"),
                    _event("blocking-schedule-save", "2026-06-01T00:01:00+00:00", mode="create"),
                ],
                "blocking-schedule-save",
            ),
            (
                "created",
                [
                    _event("user-open-the-app-for-the-first-time", "2026-06-01T00:00:00+00:00"),
                    _event("blocking-schedule-save", "2026-06-01T00:01:00+00:00", mode="create"),
                    _event("blocking-schedule-created", "2026-06-01T00:02:00+00:00"),
                ],
                "blocking-schedule-created",
            ),
        ]

        for expected_stage, events, expected_last_event in cases:
            with self.subTest(expected_stage=expected_stage):
                journey = _map_single_timeline(_build_user_timeline(events))
                self.assertEqual(journey.blocking_schedule_highest_stage, expected_stage)
                self.assertEqual(journey.last_blocking_schedule_event, expected_last_event)


def _build_mapped_journey(
    *,
    error_events: list[str] | None = None,
    error_endpoint_urls: list[str] | None = None,
    error_status_codes: list[str] | None = None,
    permission_events: list[str] | None = None,
    blocking_schedule_highest_stage: str = "not_reached",
    error_event_occurrences: list[dict[str, object]] | None = None,
) -> MappedJourney:
    """Create a compact mapped journey fixture."""
    stage_flags = {
        "pre_onboarding": True,
        "focus_bear_jr_greeting": True,
        "sign_up": True,
        "habits_introduction": False,
        "import_habits": False,
        "selecting_goals": False,
        "routine_generated": False,
        "blocking_intro": False,
        "screen_time_access": False,
        "set_up_blocking_schedule": False,
        "onboarding_complete": False,
        "home_screen": False,
    }
    return MappedJourney(
        user_id="user-123",
        distinct_id="distinct-123",
        raw_event_count=5,
        first_app_opened_at="2026-06-01T00:15:00+00:00",
        last_event_at="2026-06-01T01:15:00+00:00",
        journey_duration="1h",
        stage_flags=stage_flags,
        activation_detected=False,
        error_events=error_events or [],
        error_endpoint_urls=error_endpoint_urls or [],
        error_status_codes=error_status_codes or [],
        permission_events=permission_events or [],
        blocking_schedule_highest_stage=blocking_schedule_highest_stage,
        last_blocking_schedule_event="",
        error_event_occurrences=error_event_occurrences or [],
        top_event_counts=[],
        timeline_excerpt=[],
        llm_payload={},
    )


def _build_classified_journey(
    *,
    user_id: str,
    first_app_opened_at: str,
    last_event_at: str,
    category: str,
    error_events: list[str],
    error_endpoint_urls: list[str],
    error_status_codes: list[str],
    blocking_schedule_highest_stage: str,
    error_event_occurrences: list[dict[str, object]],
    notes: str,
    dropoff_point: str = "Routine Generated",
    onboarding_complete: str = "NO",
    pre_onboarding: str = "NO",
    sign_up: str = "NO",
) -> ClassifiedJourney:
    """Create a classified row fixture."""
    return ClassifiedJourney(
        user_id=user_id,
        first_app_opened_at=first_app_opened_at,
        last_event_at=last_event_at,
        journey_duration="3h 30m",
        category=category,
        dropoff_point=dropoff_point,
        error_events=error_events,
        error_endpoint_urls=error_endpoint_urls,
        error_status_codes=error_status_codes,
        blocking_schedule_highest_stage=blocking_schedule_highest_stage,
        error_event_occurrences=error_event_occurrences,
        notes=notes,
        activated=False,
        pre_onboarding=pre_onboarding,
        focus_bear_jr_greeting="NO",
        sign_up=sign_up,
        habits_introduction="NO",
        import_habits="NO",
        selecting_goals="NO",
        routine_generated="NO",
        blocking_intro="NO",
        screen_time_access="NO",
        set_up_blocking_schedule="NO",
        onboarding_complete=onboarding_complete,
        home_screen="NO",
        raw_event_count=12,
    )


def _build_user_timeline(events: list[dict[str, object]]) -> UserTimeline:
    """Create a compact timeline fixture for mapper tests."""
    return UserTimeline(
        user=CandidateUser(
            person_id="user-123",
            distinct_id="distinct-123",
            distinct_ids=["distinct-123"],
            name="Test User",
            email="test@example.com",
            properties={},
            raw={},
        ),
        events=events,
    )


def _event(event_name: str, timestamp: str, **properties: object) -> dict[str, object]:
    """Create a compact event fixture."""
    return {
        "id": f"{event_name}-{timestamp}",
        "event": event_name,
        "timestamp": timestamp,
        "properties": properties,
    }


def _rgb(cell: object) -> str:
    """Return an RGB-style fill value for assertions."""
    return str(cell.fill.fgColor.rgb or "")


if __name__ == "__main__":
    unittest.main()
