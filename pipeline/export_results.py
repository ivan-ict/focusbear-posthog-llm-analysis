"""Export classified journeys to a colorized Excel workbook."""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from pipeline.classify_users import ClassifiedJourney, normalize_dropoff_point
from prompts import ALLOWED_CATEGORIES


OUTPUT_COLUMNS = [
    "User ID",
    "First App Opened At",
    "Last Event At",
    "Journey Duration",
    "Category",
    "Dropoff Point",
    "Error Events",
    "Error Endpoint URLs",
    "Error Status Codes",
    "Blocking Schedule Highest Stage",
    "Notes",
    "Pre Onboarding",
    "Focus Bear Jr Greeting",
    "Sign Up",
    "Habits Introduction",
    "Import Habits",
    "Selecting Goals",
    "Routine Generated",
    "Blocking Intro",
    "Screen Time Access",
    "Set Up Blocking Schedule",
    "Onboarding Complete",
    "Home Screen",
    "Raw Event Count",
]

STATUS_COLUMNS = {
    "Pre Onboarding",
    "Focus Bear Jr Greeting",
    "Sign Up",
    "Habits Introduction",
    "Import Habits",
    "Selecting Goals",
    "Routine Generated",
    "Blocking Intro",
    "Screen Time Access",
    "Set Up Blocking Schedule",
    "Onboarding Complete",
    "Home Screen",
}

CATEGORY_COLUMN = "Category"
ERROR_EVENTS_COLUMN = "Error Events"
ERROR_ENDPOINT_URLS_COLUMN = "Error Endpoint URLs"
ERROR_STATUS_CODES_COLUMN = "Error Status Codes"
BLOCKING_SCHEDULE_HIGHEST_STAGE_COLUMN = "Blocking Schedule Highest Stage"
NOTES_COLUMN = "Notes"
DATE_COLUMNS = {"First App Opened At", "Last Event At"}
WRAP_TEXT_COLUMNS = {
    ERROR_EVENTS_COLUMN,
    ERROR_ENDPOINT_URLS_COLUMN,
    ERROR_STATUS_CODES_COLUMN,
    BLOCKING_SCHEDULE_HIGHEST_STAGE_COLUMN,
    NOTES_COLUMN,
}
MELBOURNE_TZ = ZoneInfo("Australia/Melbourne")
EXCEL_DATETIME_FORMAT = "DD/MM/YYYY HH:mm"

YES_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
NO_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
STRIPE_EVEN_FILL = PatternFill(fill_type="solid", fgColor="F7F7F7")
CATEGORY_FILLS = {
    "Permission issue": PatternFill(fill_type="solid", fgColor="FFD966"),
    "Backend issue": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "Early drop": PatternFill(fill_type="solid", fgColor="F9CB9C"),
    "Exploration without activation": PatternFill(fill_type="solid", fgColor="9FC5E8"),
    "Misclassified / already activated": PatternFill(fill_type="solid", fgColor="B6D7A8"),
}

HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
MAX_COLUMN_WIDTH = 40
MIN_TEXT_COLUMN_WIDTH = 24
SUMMARY_SHEET_TITLE = "Summary"


@dataclass(slots=True)
class AnalysisMetadata:
    """Run metadata shown on the summary sheet."""

    cohort_id: str
    cohort_name: str
    cohort_total_count: int
    analyzed_user_count: int
    posthog_user_limit: int | None
    lookback_days: int
    generated_at: datetime


def export_results(rows: list[ClassifiedJourney], output_path: Path, metadata: AnalysisMetadata) -> Path:
    """Write the final Excel output."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Onboarding Analysis"
    worksheet.freeze_panes = "A2"

    worksheet.append(OUTPUT_COLUMNS)
    _style_header_row(worksheet)

    status_column_indexes = _column_indexes(STATUS_COLUMNS)
    wrap_column_indexes = _column_indexes(WRAP_TEXT_COLUMNS)
    date_column_indexes = _column_indexes(DATE_COLUMNS)
    category_column_index = _column_index(CATEGORY_COLUMN)

    for row in rows:
        worksheet.append(_build_row_values(row))
        row_index = worksheet.max_row
        _style_record_row(
            worksheet,
            row_index=row_index,
            status_column_indexes=status_column_indexes,
            wrap_column_indexes=wrap_column_indexes,
            date_column_indexes=date_column_indexes,
            category_column_index=category_column_index,
        )

    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)
    _build_summary_sheet(workbook, rows, metadata)
    workbook.save(output_path)
    return output_path


def _build_row_values(row: ClassifiedJourney) -> list[Any]:
    """Return export values in the workbook column order."""
    error_events = ", ".join(row.error_events) if row.category == "Backend issue" else ""
    error_endpoint_urls = ", ".join(row.error_endpoint_urls)
    error_status_codes = ", ".join(row.error_status_codes)
    dropoff_point = normalize_dropoff_point(row.dropoff_point)
    return [
        row.user_id,
        _format_excel_datetime_value(row.first_app_opened_at),
        _format_excel_datetime_value(row.last_event_at),
        row.journey_duration,
        row.category,
        dropoff_point,
        error_events,
        error_endpoint_urls,
        error_status_codes,
        row.blocking_schedule_highest_stage,
        row.notes,
        row.pre_onboarding,
        row.focus_bear_jr_greeting,
        row.sign_up,
        row.habits_introduction,
        row.import_habits,
        row.selecting_goals,
        row.routine_generated,
        row.blocking_intro,
        row.screen_time_access,
        row.set_up_blocking_schedule,
        row.onboarding_complete,
        row.home_screen,
        row.raw_event_count,
    ]


def _build_summary_sheet(workbook: Workbook, rows: list[ClassifiedJourney], metadata: AnalysisMetadata) -> None:
    """Write a deterministic summary sheet for stakeholder review."""
    worksheet = workbook.create_sheet(SUMMARY_SHEET_TITLE)
    chart_ranges: dict[str, tuple[int, int]] = {}
    total_rows = len(rows)
    onboarding_completed = sum(1 for row in rows if row.onboarding_complete == "YES")
    category_counts = Counter(row.category for row in rows)
    error_event_totals = _ranked_error_event_totals(rows)

    _append_section_header(worksheet, ["Metric", "Value"])
    worksheet.append(["Generated At", _format_excel_datetime_value(metadata.generated_at.isoformat())])
    worksheet.append(["Cohort ID", metadata.cohort_id])
    worksheet.append(["Cohort Name", metadata.cohort_name or "Unknown"])
    worksheet.append(["Cohort Users Available", metadata.cohort_total_count])
    worksheet.append(["Users Analyzed", metadata.analyzed_user_count])
    worksheet.append(["Applied User Limit", _format_user_limit(metadata.posthog_user_limit)])
    worksheet.append(["Lookback Window (Days)", metadata.lookback_days])
    worksheet.append(["Onboarding Completed", onboarding_completed])
    worksheet.append(["Onboarding Completed %", _format_percentage(onboarding_completed, total_rows)])
    _style_summary_metric_values(worksheet)

    worksheet.append([])
    category_header_row = worksheet.max_row + 1
    _append_section_header(worksheet, ["Category", "Count", "Percent"])
    for category in ALLOWED_CATEGORIES:
        count = category_counts.get(category, 0)
        worksheet.append([category, count, _format_percentage(count, total_rows)])
    chart_ranges["categories"] = (category_header_row, worksheet.max_row)

    worksheet.append([])
    dropoff_header_row = worksheet.max_row + 1
    _append_section_header(worksheet, ["Top Dropoff Point", "Count"])
    for label, count in _ranked_dropoff_counts(rows):
        worksheet.append([label, count])
    chart_ranges["dropoffs"] = (dropoff_header_row, worksheet.max_row)

    worksheet.append([])
    _append_section_header(worksheet, ["Blocking Schedule Deepest Stage", "Count", "Percent"])
    for stage_name, count in _ranked_blocking_schedule_highest_stage_counts(rows):
        worksheet.append([stage_name, count, _format_percentage(count, total_rows)])

    worksheet.append([])
    _append_section_header(worksheet, ["Error Event Totals"])
    worksheet.append(["Event", "Raw Events", "Affected Users", "Affected Users %"])
    _style_summary_subheader_row(worksheet)
    error_totals_header_row = worksheet.max_row
    for event_name, raw_events, affected_users in error_event_totals:
        worksheet.append(
            [
                event_name,
                raw_events,
                affected_users,
                _format_percentage(affected_users, total_rows),
            ]
        )
    chart_ranges["error_events"] = (error_totals_header_row, worksheet.max_row)

    worksheet.append([])
    _append_section_header(worksheet, ["Error Breakdown"])
    worksheet.append(["Event", "Endpoint URL", "Status Code", "Raw Events", "Affected Users", "Affected Users %"])
    _style_summary_subheader_row(worksheet)
    for event_name, endpoint_url, status_code, raw_events, affected_users in _ranked_error_breakdown_rows(rows):
        worksheet.append(
            [
                event_name,
                endpoint_url,
                status_code,
                raw_events,
                affected_users,
                _format_percentage(affected_users, total_rows),
            ]
        )

    worksheet.append([])
    _append_section_header(worksheet, ["Key Finding"])
    for finding in _build_key_findings(rows, category_counts):
        worksheet.append([finding])

    _add_summary_charts(worksheet, chart_ranges)
    _autosize_columns(worksheet)


def _append_section_header(worksheet: Worksheet, values: list[str]) -> None:
    """Append and style a summary section header row."""
    worksheet.append(values)
    row_index = worksheet.max_row
    for cell in worksheet[row_index]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT if len(values) > 1 else WRAP_ALIGNMENT


def _style_summary_subheader_row(worksheet: Worksheet) -> None:
    """Style the row immediately appended after a single-cell section header."""
    row_index = worksheet.max_row
    for cell in worksheet[row_index]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT


def _style_summary_metric_values(worksheet: Worksheet) -> None:
    """Apply consistent alignment/formatting to metric values."""
    for row_index in range(2, worksheet.max_row + 1):
        label_cell = worksheet.cell(row=row_index, column=1)
        value_cell = worksheet.cell(row=row_index, column=2)
        if not str(label_cell.value or "").strip():
            continue
        if isinstance(value_cell.value, datetime):
            value_cell.number_format = EXCEL_DATETIME_FORMAT
        value_cell.alignment = WRAP_ALIGNMENT


def _ranked_dropoff_counts(rows: list[ClassifiedJourney]) -> list[tuple[str, int]]:
    """Return dropoff counts sorted by frequency, excluding unknown values."""
    counts = Counter(normalize_dropoff_point(row.dropoff_point) for row in rows)
    counts.pop("Unknown", None)
    if not counts:
        return [("None", 0)]
    return counts.most_common()


def _ranked_blocking_schedule_highest_stage_counts(
    rows: list[ClassifiedJourney],
) -> list[tuple[str, int]]:
    """Return deepest blocking-schedule stages ranked by frequency."""
    counts = Counter(row.blocking_schedule_highest_stage for row in rows)
    counts.pop("not_reached", None)
    if not counts:
        return [("None", 0)]
    return counts.most_common()


def _ranked_error_event_totals(rows: list[ClassifiedJourney]) -> list[tuple[str, int, int]]:
    """Return canonical error events ranked by affected users and raw count."""
    raw_counts, affected_users = _aggregate_error_breakdown(rows)
    event_rows: list[tuple[str, int, int]] = []
    by_event: dict[str, tuple[int, set[str]]] = {}
    for (event_name, _endpoint_url, _status_code), raw_count in raw_counts.items():
        current_raw, current_users = by_event.get(event_name, (0, set()))
        current_users = set(current_users)
        current_raw += raw_count
        current_users.update(affected_users.get((event_name, _endpoint_url, _status_code), set()))
        by_event[event_name] = (current_raw, current_users)

    for event_name, (raw_count, users) in by_event.items():
        event_rows.append((event_name, raw_count, len(users)))

    if not event_rows:
        return [("None", 0, 0)]

    return sorted(event_rows, key=lambda item: (-item[2], -item[1], item[0]))


def _ranked_error_breakdown_rows(
    rows: list[ClassifiedJourney],
) -> list[tuple[str, str, str, int, int]]:
    """Return canonical error breakdown rows keyed by event, endpoint, and status."""
    raw_counts, affected_users = _aggregate_error_breakdown(rows)
    if not raw_counts:
        return [("None", "None", "None", 0, 0)]

    def _sort_key(item: tuple[tuple[str, str, str], int]) -> tuple[int, int, str, str, str]:
        key, raw_count = item
        return (-len(affected_users.get(key, set())), -raw_count, key[0], key[1], key[2])

    ranked_rows: list[tuple[str, str, str, int, int]] = []
    for key, raw_count in sorted(raw_counts.items(), key=_sort_key):
        event_name, endpoint_url, status_code = key
        ranked_rows.append(
            (
                event_name or "None",
                endpoint_url or "(missing)",
                status_code or "(missing)",
                raw_count,
                len(affected_users.get(key, set())),
            )
        )
    return ranked_rows


def _aggregate_error_breakdown(
    rows: list[ClassifiedJourney],
) -> tuple[Counter[tuple[str, str, str]], dict[tuple[str, str, str], set[str]]]:
    """Aggregate canonical error tuples into raw-event and affected-user counts."""
    raw_counts: Counter[tuple[str, str, str]] = Counter()
    affected_users: dict[tuple[str, str, str], set[str]] = {}
    for row in rows:
        for occurrence in row.error_event_occurrences:
            key = (
                str(occurrence.get("event") or "").strip(),
                str(occurrence.get("endpoint_url") or "").strip(),
                str(occurrence.get("status_code") or "").strip(),
            )
            raw_counts[key] += int(occurrence.get("count") or 0)
            affected_users.setdefault(key, set()).add(row.user_id)
    return raw_counts, affected_users


def _build_key_findings(
    rows: list[ClassifiedJourney],
    category_counts: Counter[str],
) -> list[str]:
    """Return short deterministic headline findings from the workbook rows."""
    total_rows = len(rows)
    if total_rows == 0:
        return ["No users were analyzed."]

    findings: list[str] = []

    top_category, top_category_count = max(
        ((category, category_counts.get(category, 0)) for category in ALLOWED_CATEGORIES),
        key=lambda item: (item[1], item[0]),
    )
    findings.append(
        f"Largest category: {top_category} ({top_category_count}/{total_rows}, "
        f"{_format_percentage(top_category_count, total_rows)})."
    )

    top_dropoff, top_dropoff_count = _ranked_dropoff_counts(rows)[0]
    if top_dropoff != "None":
        findings.append(f"Most common dropoff point: {top_dropoff} ({top_dropoff_count} users).")
    else:
        findings.append("No dropoff point could be determined from the analyzed users.")

    top_error_event, top_error_raw_count, top_error_affected_users = _ranked_error_event_totals(rows)[0]
    if top_error_event != "None":
        findings.append(
            "Most common canonical error event: "
            f"{top_error_event} ({top_error_raw_count} raw events across {top_error_affected_users} users)."
        )
    else:
        findings.append("No canonical error events were recorded in the analyzed users.")

    top_error_breakdown = _ranked_error_breakdown_rows(rows)[0]
    if top_error_breakdown[0] != "None":
        findings.append(
            "Most affected error breakdown: "
            f"{top_error_breakdown[0]} @ {top_error_breakdown[1]} "
            f"[{top_error_breakdown[2]}] ({top_error_breakdown[4]} users)."
        )
    else:
        findings.append("No error breakdown rows were recorded in the analyzed users.")

    top_blocking_stage, top_blocking_stage_count = _ranked_blocking_schedule_highest_stage_counts(rows)[0]
    if top_blocking_stage != "None":
        findings.append(
            f"Most common blocking-schedule deepest stage: {top_blocking_stage} "
            f"({top_blocking_stage_count} users)."
        )
    else:
        findings.append("No users reached blocking schedule in the analyzed users.")

    onboarding_completed = sum(1 for row in rows if row.onboarding_complete == "YES")
    findings.append(
        f"Onboarding completion rate: {_format_percentage(onboarding_completed, total_rows)} "
        f"({onboarding_completed}/{total_rows})."
    )
    return findings


def _add_summary_charts(worksheet: Worksheet, chart_ranges: dict[str, tuple[int, int]]) -> None:
    """Add a compact set of native Excel charts to the summary sheet."""
    _add_bar_chart(
        worksheet,
        title="Journey Categories",
        anchor="H2",
        header_row=chart_ranges["categories"][0],
        end_row=chart_ranges["categories"][1],
        category_col=1,
        value_col=2,
    )
    _add_bar_chart(
        worksheet,
        title="Top Dropoff Points",
        anchor="H20",
        header_row=chart_ranges["dropoffs"][0],
        end_row=chart_ranges["dropoffs"][1],
        category_col=1,
        value_col=2,
    )
    _add_bar_chart(
        worksheet,
        title="Error Events by Affected Users",
        anchor="H38",
        header_row=chart_ranges["error_events"][0],
        end_row=chart_ranges["error_events"][1],
        category_col=1,
        value_col=3,
    )


def _add_bar_chart(
    worksheet: Worksheet,
    *,
    title: str,
    anchor: str,
    header_row: int,
    end_row: int,
    category_col: int,
    value_col: int,
) -> None:
    """Render a single horizontal bar chart from a summary table."""
    if end_row <= header_row:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.height = 7
    chart.width = 12
    chart.legend = None

    data = Reference(
        worksheet,
        min_col=value_col,
        max_col=value_col,
        min_row=header_row,
        max_row=end_row,
    )
    categories = Reference(
        worksheet,
        min_col=category_col,
        max_col=category_col,
        min_row=header_row + 1,
        max_row=end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, anchor)


def _format_user_limit(limit: int | None) -> str:
    """Return a stable summary label for the user-limit setting."""
    if limit is None:
        return "All cohort users"
    return str(limit)


def _format_percentage(count: int, total: int) -> str:
    """Return a stable percentage string for summary output."""
    if total <= 0:
        return "0.0%"
    return f"{(count / total) * 100:.1f}%"


def _format_excel_datetime_value(value: str) -> datetime | str:
    """Convert an ISO timestamp to a Melbourne-local naive datetime for Excel."""
    if not value:
        return ""

    normalized = str(value).strip().replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return value

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=MELBOURNE_TZ)
    else:
        parsed = parsed.astimezone(MELBOURNE_TZ)
    return parsed.replace(tzinfo=None)


def _style_header_row(worksheet: Worksheet) -> None:
    """Apply workbook styling to the header row."""
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT


def _style_record_row(
    worksheet: Worksheet,
    row_index: int,
    status_column_indexes: set[int],
    wrap_column_indexes: set[int],
    date_column_indexes: set[int],
    category_column_index: int,
) -> None:
    """Apply row striping, wrapping, category fills, and status styling."""
    _apply_row_striping(worksheet, row_index)
    _style_category_cell(worksheet.cell(row=row_index, column=category_column_index))

    for column_index in wrap_column_indexes:
        worksheet.cell(row=row_index, column=column_index).alignment = WRAP_ALIGNMENT

    for column_index in status_column_indexes:
        cell = worksheet.cell(row=row_index, column=column_index)
        cell.alignment = CENTER_ALIGNMENT
        _style_status_cell(cell)

    for column_index in date_column_indexes:
        cell = worksheet.cell(row=row_index, column=column_index)
        if isinstance(cell.value, datetime):
            cell.number_format = EXCEL_DATETIME_FORMAT


def _apply_row_striping(worksheet: Worksheet, row_index: int) -> None:
    """Apply subtle alternating fill across non-header rows."""
    if row_index % 2 != 0:
        return
    for cell in worksheet[row_index]:
        cell.fill = STRIPE_EVEN_FILL


def _style_category_cell(cell: Any) -> None:
    """Color the category cell using a fixed mapping."""
    fill = CATEGORY_FILLS.get(str(cell.value or "").strip())
    if fill is not None:
        cell.fill = fill


def _style_status_cell(cell: Any) -> None:
    """Color YES and NO cells in the status columns."""
    cell_value = str(cell.value or "").strip().upper()
    if cell_value == "YES":
        cell.fill = YES_FILL
    elif cell_value == "NO":
        cell.fill = NO_FILL


def _column_index(column_name: str) -> int:
    """Return the 1-based index of a workbook column."""
    return OUTPUT_COLUMNS.index(column_name) + 1


def _column_indexes(column_names: set[str]) -> set[int]:
    """Return 1-based indexes for the provided workbook columns."""
    return {
        index
        for index, column_name in enumerate(OUTPUT_COLUMNS, start=1)
        if column_name in column_names
    }


def _autosize_columns(worksheet: Worksheet) -> None:
    """Resize columns to a readable width with a fixed cap."""
    for column in worksheet.columns:
        values = [str(cell.value or "") for cell in column]
        max_length = max((len(value) for value in values), default=0)
        column_letter = column[0].column_letter
        column_name = str(column[0].value or "")
        width = min(max_length + 2, MAX_COLUMN_WIDTH)
        if column_name in WRAP_TEXT_COLUMNS:
            width = max(width, MIN_TEXT_COLUMN_WIDTH)
        worksheet.column_dimensions[column_letter].width = width
