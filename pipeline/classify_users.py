"""Classify mapped onboarding journeys with OpenAI."""

from __future__ import annotations

from dataclasses import asdict, dataclass
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from config import AppConfig
from clients.openai_client import OpenAIClient
from pipeline.map_events import MappedJourney, STAGE_LABELS
from prompts import ALLOWED_CATEGORIES, REQUIRED_OUTPUT_KEYS, STAGE_KEYS


@dataclass(slots=True)
class ClassifiedJourney:
    """Final normalized output row written to Excel."""

    user_id: str
    first_app_opened_at: str
    last_event_at: str
    journey_duration: str
    category: str
    dropoff_point: str
    error_events: list[str]
    error_endpoint_urls: list[str]
    error_status_codes: list[str]
    blocking_schedule_highest_stage: str
    error_event_occurrences: list[dict[str, Any]]
    notes: str
    activated: bool
    pre_onboarding: str
    focus_bear_jr_greeting: str
    sign_up: str
    habits_introduction: str
    import_habits: str
    selecting_goals: str
    routine_generated: str
    blocking_intro: str
    screen_time_access: str
    set_up_blocking_schedule: str
    onboarding_complete: str
    home_screen: str
    raw_event_count: int


def classify_users(
    config: AppConfig,
    journeys: list[MappedJourney],
    openai_client: OpenAIClient | None = None,
) -> list[ClassifiedJourney]:
    """Classify all journeys using the configured source."""
    if config.classification_source == "cached":
        return _load_cached_or_bootstrap(config=config, journeys=journeys)
    if config.classification_source == "fallback":
        results = [_fallback_classification(journey, "CLASSIFICATION_SOURCE=fallback") for journey in journeys]
        _write_classification_cache(config.classified_journeys_cache_path, results)
        return results
    if openai_client is None:
        raise ValueError("OpenAI client is required when CLASSIFICATION_SOURCE=openai.")

    results: list[ClassifiedJourney] = []
    for journey in journeys:
        print(f"Classifying user {journey.user_id}...", flush=True)
        try:
            raw_response = openai_client.classify_user(journey.llm_payload)
            results.append(_normalize_response(journey, raw_response))
        except Exception as exc:
            print(f"Classification failed for user {journey.user_id}: {exc}", flush=True)
            results.append(_fallback_classification(journey, str(exc)))
    _write_classification_cache(config.classified_journeys_cache_path, results)
    return results


def _load_cached_or_bootstrap(config: AppConfig, journeys: list[MappedJourney]) -> list[ClassifiedJourney]:
    """Load cached classifications or bootstrap them from the current workbook."""
    if config.classified_journeys_cache_path.exists():
        return _load_cached_classifications(config.classified_journeys_cache_path, journeys)
    if config.output_xlsx_path.exists():
        results = _load_classifications_from_workbook(config.output_xlsx_path, journeys)
        _write_classification_cache(config.classified_journeys_cache_path, results)
        return results
    raise FileNotFoundError(
        "No cached classifications were found. Expected either "
        f"{config.classified_journeys_cache_path} or {config.output_xlsx_path}. "
        "Run once with CLASSIFICATION_SOURCE=openai to refresh local results."
    )


def _load_cached_classifications(
    cache_path: Path,
    journeys: list[MappedJourney],
) -> list[ClassifiedJourney]:
    """Load cached classifications from JSON and validate they match the current journeys."""
    with cache_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, list):
        raise ValueError(f"Expected a JSON list at {cache_path}")
    results = [_classified_journey_from_dict(item, journeys) for item in payload]
    _ensure_cached_users_match(journeys, results, source_path=cache_path)
    return results


def _load_classifications_from_workbook(
    workbook_path: Path,
    journeys: list[MappedJourney],
) -> list[ClassifiedJourney]:
    """Rebuild cached classifications by combining workbook labels with current mapped journeys."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook["Onboarding Analysis"]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"No rows found in workbook: {workbook_path}")

    header = [str(value or "").strip() for value in rows[0]]
    workbook_by_user: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        row_values = dict(zip(header, row))
        user_id = str(row_values.get("User ID") or "").strip()
        if user_id:
            workbook_by_user[user_id] = row_values

    results = [_classified_journey_from_workbook_row(journey, workbook_by_user) for journey in journeys]
    _ensure_cached_users_match(journeys, results, source_path=workbook_path)
    return results


def _classified_journey_from_workbook_row(
    journey: MappedJourney,
    workbook_by_user: dict[str, dict[str, Any]],
) -> ClassifiedJourney:
    """Build a classified journey from workbook-only fields plus deterministic mapped data."""
    row = workbook_by_user.get(journey.user_id)
    if row is None:
        raise ValueError(
            f"Workbook cache is missing user {journey.user_id}. "
            "Run once with CLASSIFICATION_SOURCE=openai to refresh local results."
        )
    return ClassifiedJourney(
        user_id=journey.user_id,
        first_app_opened_at=journey.first_app_opened_at,
        last_event_at=journey.last_event_at,
        journey_duration=journey.journey_duration,
        category=_normalize_category(row.get("Category")),
        dropoff_point=normalize_dropoff_point(row.get("Dropoff Point")),
        error_events=journey.error_events,
        error_endpoint_urls=journey.error_endpoint_urls,
        error_status_codes=journey.error_status_codes,
        blocking_schedule_highest_stage=journey.blocking_schedule_highest_stage,
        error_event_occurrences=journey.error_event_occurrences,
        notes=str(row.get("Notes") or "").strip() or "No note returned by the model.",
        activated=_excel_yes_no(row.get("Onboarding Complete")) == "YES" or journey.activation_detected,
        pre_onboarding=_excel_yes_no(row.get("Pre Onboarding")),
        focus_bear_jr_greeting=_excel_yes_no(row.get("Focus Bear Jr Greeting")),
        sign_up=_excel_yes_no(row.get("Sign Up")),
        habits_introduction=_excel_yes_no(row.get("Habits Introduction")),
        import_habits=_excel_yes_no(row.get("Import Habits")),
        selecting_goals=_excel_yes_no(row.get("Selecting Goals")),
        routine_generated=_excel_yes_no(row.get("Routine Generated")),
        blocking_intro=_excel_yes_no(row.get("Blocking Intro")),
        screen_time_access=_excel_yes_no(row.get("Screen Time Access")),
        set_up_blocking_schedule=_excel_yes_no(row.get("Set Up Blocking Schedule")),
        onboarding_complete=_excel_yes_no(row.get("Onboarding Complete")),
        home_screen=_excel_yes_no(row.get("Home Screen")),
        raw_event_count=journey.raw_event_count,
    )


def _classified_journey_from_dict(payload: dict[str, Any], journeys: list[MappedJourney]) -> ClassifiedJourney:
    """Convert a cached dict into a classified journey with deterministic fields refreshed."""
    journey_by_user = {journey.user_id: journey for journey in journeys}
    user_id = str(payload.get("user_id") or "").strip()
    journey = journey_by_user.get(user_id)
    if journey is None:
        raise ValueError(
            f"Cached classifications include user {user_id or '(missing)'}, "
            "which does not match the current local raw dataset."
        )
    return ClassifiedJourney(
        user_id=journey.user_id,
        first_app_opened_at=journey.first_app_opened_at,
        last_event_at=journey.last_event_at,
        journey_duration=journey.journey_duration,
        category=_normalize_category(payload.get("category")),
        dropoff_point=normalize_dropoff_point(payload.get("dropoff_point")),
        error_events=journey.error_events,
        error_endpoint_urls=journey.error_endpoint_urls,
        error_status_codes=journey.error_status_codes,
        blocking_schedule_highest_stage=journey.blocking_schedule_highest_stage,
        error_event_occurrences=journey.error_event_occurrences,
        notes=str(payload.get("notes") or "").strip() or "No note returned by the model.",
        activated=_normalize_bool(payload.get("activated")),
        pre_onboarding=_normalize_yes_no(payload.get("pre_onboarding")),
        focus_bear_jr_greeting=_normalize_yes_no(payload.get("focus_bear_jr_greeting")),
        sign_up=_normalize_yes_no(payload.get("sign_up")),
        habits_introduction=_normalize_yes_no(payload.get("habits_introduction")),
        import_habits=_normalize_yes_no(payload.get("import_habits")),
        selecting_goals=_normalize_yes_no(payload.get("selecting_goals")),
        routine_generated=_normalize_yes_no(payload.get("routine_generated")),
        blocking_intro=_normalize_yes_no(payload.get("blocking_intro")),
        screen_time_access=_normalize_yes_no(payload.get("screen_time_access")),
        set_up_blocking_schedule=_normalize_yes_no(payload.get("set_up_blocking_schedule")),
        onboarding_complete=_normalize_yes_no(payload.get("onboarding_complete")),
        home_screen=_normalize_yes_no(payload.get("home_screen")),
        raw_event_count=journey.raw_event_count,
    )


def _write_classification_cache(cache_path: Path, rows: list[ClassifiedJourney]) -> None:
    """Persist classified journeys for later reuse."""
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    payload = [asdict(row) for row in rows]
    with cache_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=True)


def _ensure_cached_users_match(
    journeys: list[MappedJourney],
    results: list[ClassifiedJourney],
    *,
    source_path: Path,
) -> None:
    """Ensure cached classifications exactly match the current local raw dataset."""
    expected_user_ids = {journey.user_id for journey in journeys}
    actual_user_ids = {row.user_id for row in results}
    if expected_user_ids == actual_user_ids:
        return

    missing = sorted(expected_user_ids - actual_user_ids)
    extra = sorted(actual_user_ids - expected_user_ids)
    mismatch_parts = []
    if missing:
        mismatch_parts.append(f"missing users: {', '.join(missing[:5])}")
    if extra:
        mismatch_parts.append(f"unexpected users: {', '.join(extra[:5])}")
    mismatch_text = "; ".join(mismatch_parts)
    raise ValueError(
        f"Cached classifications at {source_path} do not match the current local raw dataset ({mismatch_text}). "
        "Run once with CLASSIFICATION_SOURCE=openai to refresh local results."
    )


def _normalize_response(journey: MappedJourney, response: dict[str, Any]) -> ClassifiedJourney:
    """Validate and normalize the model response."""
    missing_keys = [key for key in REQUIRED_OUTPUT_KEYS if key not in response]
    if missing_keys:
        raise ValueError(f"Missing keys in OpenAI response: {', '.join(missing_keys)}")

    category = _normalize_category(response.get("category"))
    activated = _normalize_bool(response.get("activated"))
    stage_values = {stage: _normalize_yes_no(response.get(stage)) for stage in STAGE_KEYS}
    notes = str(response.get("notes") or "").strip()
    if not notes:
        notes = "No note returned by the model."
    dropoff_point = normalize_dropoff_point(response.get("dropoff_point"))

    return ClassifiedJourney(
        user_id=journey.user_id,
        first_app_opened_at=journey.first_app_opened_at,
        last_event_at=journey.last_event_at,
        journey_duration=journey.journey_duration,
        category=category,
        dropoff_point=dropoff_point,
        error_events=journey.error_events,
        error_endpoint_urls=journey.error_endpoint_urls,
        error_status_codes=journey.error_status_codes,
        blocking_schedule_highest_stage=journey.blocking_schedule_highest_stage,
        error_event_occurrences=journey.error_event_occurrences,
        notes=notes,
        activated=activated,
        pre_onboarding=stage_values["pre_onboarding"],
        focus_bear_jr_greeting=stage_values["focus_bear_jr_greeting"],
        sign_up=stage_values["sign_up"],
        habits_introduction=stage_values["habits_introduction"],
        import_habits=stage_values["import_habits"],
        selecting_goals=stage_values["selecting_goals"],
        routine_generated=stage_values["routine_generated"],
        blocking_intro=stage_values["blocking_intro"],
        screen_time_access=stage_values["screen_time_access"],
        set_up_blocking_schedule=stage_values["set_up_blocking_schedule"],
        onboarding_complete=stage_values["onboarding_complete"],
        home_screen=stage_values["home_screen"],
        raw_event_count=journey.raw_event_count,
    )


def _fallback_classification(journey: MappedJourney, error_message: str) -> ClassifiedJourney:
    """Build a deterministic fallback when the LLM response is unavailable."""
    category = "Early drop"
    if journey.activation_detected:
        category = "Misclassified / already activated"
    elif journey.error_events:
        category = "Backend issue"
    elif journey.permission_events:
        category = "Permission issue"
    elif journey.raw_event_count > 8:
        category = "Exploration without activation"

    highest_stage = _highest_stage_label(journey.stage_flags)
    notes = (
        "Fallback classification used. "
        f"Reason: {error_message}. "
        f"Activation={journey.activation_detected}. "
        f"Errors={', '.join(journey.error_events) or 'none'}. "
        f"Permissions={', '.join(journey.permission_events) or 'none'}."
    )

    return ClassifiedJourney(
        user_id=journey.user_id,
        first_app_opened_at=journey.first_app_opened_at,
        last_event_at=journey.last_event_at,
        journey_duration=journey.journey_duration,
        category=category,
        dropoff_point=highest_stage,
        error_events=journey.error_events,
        error_endpoint_urls=journey.error_endpoint_urls,
        error_status_codes=journey.error_status_codes,
        blocking_schedule_highest_stage=journey.blocking_schedule_highest_stage,
        error_event_occurrences=journey.error_event_occurrences,
        notes=notes,
        activated=journey.activation_detected,
        pre_onboarding=_yes_no(journey.stage_flags.get("pre_onboarding", False)),
        focus_bear_jr_greeting=_yes_no(journey.stage_flags.get("focus_bear_jr_greeting", False)),
        sign_up=_yes_no(journey.stage_flags.get("sign_up", False)),
        habits_introduction=_yes_no(journey.stage_flags.get("habits_introduction", False)),
        import_habits=_yes_no(journey.stage_flags.get("import_habits", False)),
        selecting_goals=_yes_no(journey.stage_flags.get("selecting_goals", False)),
        routine_generated=_yes_no(journey.stage_flags.get("routine_generated", False)),
        blocking_intro=_yes_no(journey.stage_flags.get("blocking_intro", False)),
        screen_time_access=_yes_no(journey.stage_flags.get("screen_time_access", False)),
        set_up_blocking_schedule=_yes_no(journey.stage_flags.get("set_up_blocking_schedule", False)),
        onboarding_complete=_yes_no(journey.stage_flags.get("onboarding_complete", False)),
        home_screen=_yes_no(journey.stage_flags.get("home_screen", False)),
        raw_event_count=journey.raw_event_count,
    )


def _normalize_category(value: Any) -> str:
    """Normalize a model category to one of the allowed values."""
    text = str(value or "").strip()
    for allowed in ALLOWED_CATEGORIES:
        if text.lower() == allowed.lower():
            return allowed
    raise ValueError(f"Invalid category returned by model: {text}")


def _normalize_bool(value: Any) -> bool:
    """Normalize booleans returned by the model."""
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if text in {"true", "yes", "1"}:
        return True
    if text in {"false", "no", "0"}:
        return False
    raise ValueError(f"Invalid activated value returned by model: {value}")


def _normalize_yes_no(value: Any) -> str:
    """Normalize a model field to YES or NO."""
    if isinstance(value, bool):
        return "YES" if value else "NO"
    text = str(value or "").strip().upper()
    if text in {"YES", "NO"}:
        return text
    raise ValueError(f"Invalid YES/NO value returned by model: {value}")


def _excel_yes_no(value: Any) -> str:
    """Normalize a workbook cell value to YES/NO."""
    if value is None or str(value).strip() == "":
        return "NO"
    return _normalize_yes_no(value)


def _highest_stage_label(stage_flags: dict[str, bool]) -> str:
    """Return the furthest stage reached from the deterministic mapping."""
    reached = [STAGE_LABELS[stage] for stage in STAGE_KEYS if stage_flags.get(stage)]
    return reached[-1] if reached else "Unknown"


def normalize_dropoff_point(value: Any) -> str:
    """Normalize a model dropoff point to a canonical stage label."""
    text = str(value or "").strip()
    if not text:
        return "Unknown"

    normalized = _normalize_dropoff_key(text)
    if normalized == "unknown":
        return "Unknown"

    for stage in STAGE_KEYS:
        stage_label = STAGE_LABELS[stage]
        if normalized in {_normalize_dropoff_key(stage), _normalize_dropoff_key(stage_label)}:
            return stage_label

    return "Unknown"


def _normalize_dropoff_key(value: str) -> str:
    """Return a stable comparison key for dropoff point variants."""
    collapsed = re.sub(r"[_\-\s]+", " ", value.strip().lower())
    return collapsed.strip()


def _yes_no(value: bool) -> str:
    """Return YES or NO for a boolean."""
    return "YES" if value else "NO"
